import frappe
from frappe.utils import date_diff
import json
from io import BytesIO
import openpyxl
from openpyxl.styles import PatternFill, Font

def execute(filters=None):
    columns = get_columns()
    data = get_data(filters)
    return columns, data

def get_columns():
    return [
        {"fieldname": "sales_order", "label": "Sales Order", "fieldtype": "Link", "options": "Sales Order", "width": 130},
        {"fieldname": "so_date", "label": "SO Date", "fieldtype": "Date", "width": 100},
        {"fieldname": "customer", "label": "Customer ID", "fieldtype": "Link", "options": "Customer", "width": 130},
        {"fieldname": "customer_name", "label": "Customer Name", "fieldtype": "Data", "width": 180},
        {"fieldname": "customer_add", "label": "Customer Address", "fieldtype": "Data", "width": 200},
        {"fieldname": "shipping_address", "label": "Shipping Address", "fieldtype": "Data", "width": 200},
        {"fieldname": "created_by", "label": "Created By", "fieldtype": "Data", "width": 140},
        {"fieldname": "item_code", "label": "Item Code", "fieldtype": "Link", "options": "Item", "width": 130},
        {"fieldname": "item_name", "label": "Item Name", "fieldtype": "Data", "width": 200},
        {"fieldname": "sale_qty", "label": "Sale Qty", "fieldtype": "Float", "width": 100},
        {"fieldname": "delivered_qty", "label": "Delivered Qty", "fieldtype": "Float", "width": 100},
        {"fieldname": "pending_qty", "label": "Pending Qty", "fieldtype": "Float", "width": 100},
        {"fieldname": "sale_uom", "label": "Sale UOM", "fieldtype": "Link", "options": "UOM", "width": 100},
        {"fieldname": "remarks", "label": "Remarks", "fieldtype": "Data", "width": 200},
        {"fieldname": "billed_amount", "label": "Bill Amount", "fieldtype": "Currency", "width": 120},
        {"fieldname": "outstanding_amount", "label": "Outstanding Amount", "fieldtype": "Currency", "width": 140},
        {"fieldname": "delay_dispatch", "label": "Delay to Dispatch (Days)", "fieldtype": "Int", "width": 170},
        {"fieldname": "delay_status", "label": "Delay Status", "fieldtype": "Data", "width": 120}
    ]

def get_data(filters):
    conditions = []
    if filters and filters.get("from_date") and filters.get("to_date"):
        conditions.append(f"so.transaction_date BETWEEN '{filters.get('from_date')}' AND '{filters.get('to_date')}'")
    if filters and filters.get("customer"):
        conditions.append(f"so.customer = '{filters.get('customer')}'")
        
    condition_str = ""
    if conditions:
        condition_str = "WHERE " + " AND ".join(conditions)
    
    # New Item-Level SQL to pull granular tracking
    sql = f"""
        SELECT
            so.name as sales_order,
            so.transaction_date as so_date,
            so.customer,
            so.customer_name,
            so.customer_address as customer_add,
            so.shipping_address_name as shipping_address,
            so.owner as created_by,
            soi.description as remarks,
            soi.item_code,
            soi.item_name,
            soi.qty as sale_qty,
            soi.delivered_qty as delivered_qty,
            (soi.qty - soi.delivered_qty) as pending_qty,
            soi.uom as sale_uom,
            soi.billed_amt as billed_amount,
            
            -- Fetch outstanding amount at SO level using a subquery
            (SELECT SUM(outstanding_amount) 
             FROM `tabSales Invoice` 
             WHERE docstatus = 1 AND name IN 
                 (SELECT parent FROM `tabSales Invoice Item` WHERE sales_order = so.name)
            ) as outstanding_amount,
            
            -- Fetch latest delivery date for this specific item using a subquery
            (SELECT MAX(dn.posting_date) 
             FROM `tabDelivery Note` dn 
             JOIN `tabDelivery Note Item` dni ON dn.name = dni.parent 
             WHERE dni.so_detail = soi.name AND dn.docstatus = 1
            ) as latest_dn_date
            
        FROM `tabSales Order Item` soi
        JOIN `tabSales Order` so ON so.name = soi.parent
        {condition_str}
        ORDER BY so.transaction_date DESC, so.name ASC
    """
    
    raw_data = frappe.db.sql(sql, as_dict=True)
    
    data = []
    for row in raw_data:
        delay_dispatch = None
        delay_status = ""
        
        # Calculate delay based on the latest DN date vs SO date
        if row.so_date and row.latest_dn_date:
            delay_dispatch = date_diff(row.latest_dn_date, row.so_date)
            if delay_dispatch <= 3:
                delay_status = "On Time"
            elif 3 < delay_dispatch <= 5:
                delay_status = "Warning"
            else:
                delay_status = "Delayed"
                
        row.update({
            "delay_dispatch": delay_dispatch,
            "delay_status": delay_status
        })
        data.append(row)
        
    return data

@frappe.whitelist()
def download_colored_excel(filters=None):
    if isinstance(filters, str):
        filters = json.loads(filters)
        
    filters = frappe._dict(filters or {})
        
    columns = get_columns()
    data = get_data(filters)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Movement Summary"
    
    # Write Headers
    for col_idx, col in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col.get("label"))
        cell.font = Font(bold=True)
        
    # Colors matching JS logic
    green_fill = PatternFill(start_color="e7f5ee", end_color="e7f5ee", fill_type="solid")
    green_font = Font(color="1e8a5b")
    
    yellow_fill = PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid")
    yellow_font = Font(color="856404")
    
    red_fill = PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid")
    red_font = Font(color="721c24")
    
    for row_idx, row in enumerate(data, start=2):
        delay = row.get("delay_dispatch")
        fill_to_apply = None
        font_to_apply = None
        
        if delay is not None:
            if delay <= 3:
                fill_to_apply = green_fill
                font_to_apply = green_font
            elif 3 < delay <= 5:
                fill_to_apply = yellow_fill
                font_to_apply = yellow_font
            elif delay > 5:
                fill_to_apply = red_fill
                font_to_apply = red_font
                
        for col_idx, col in enumerate(columns, start=1):
            fieldname = col.get("fieldname")
            val = row.get(fieldname)
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            
            if fill_to_apply and font_to_apply:
                cell.fill = fill_to_apply
                cell.font = font_to_apply
                
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Output to response
    file_data = BytesIO()
    wb.save(file_data)
    
    frappe.response['filename'] = "Sales_Movement_Summary.xlsx"
    frappe.response['filecontent'] = file_data.getvalue()
    frappe.response['type'] = 'binary'
