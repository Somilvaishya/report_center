# Copyright (c) 2026, Antigravity and contributors
# For license information, please see license.txt

import frappe
from frappe import _
from erpnext.accounts.report.general_ledger.general_ledger import execute as execute_gl
import json
from io import BytesIO
import openpyxl
from openpyxl.styles import PatternFill, Font


def execute(filters=None):
	"""
	General Ledger Audit Report
	Reuses standard General Ledger execution logic and appends custom columns.
	"""
	# Call the standard ERPNext General Ledger report logic
	columns, data = execute_gl(filters)

	# Locate 'posting_date' to insert 'Created On' column right after it
	insert_index = 2  # default fallback index
	for i, col in enumerate(columns):
		if col.get("fieldname") == "posting_date":
			insert_index = i + 1
			break

	# Insert the 'Created On' column
	columns.insert(insert_index, {
		"label": _("Created On"),
		"fieldname": "creation",
		"fieldtype": "Datetime",
		"width": 150
	})
	
	# Append the new GST and Supplier columns
	columns.extend([
		{"label": _("Company GSTIN"), "fieldname": "company_gstin", "fieldtype": "Data", "width": 140},
		{"label": _("Booking Location of the company"), "fieldname": "billing_address_display", "fieldtype": "Small Text", "width": 200},
		{"label": _("Supplier Billing GSTIN"), "fieldname": "supplier_billing_gstin", "fieldtype": "Data", "width": 140},
		{"label": _("Supplier Shipping GSTIN"), "fieldname": "supplier_shipping_gstin", "fieldtype": "Data", "width": 140},
		{"label": _("Supplier Billing State"), "fieldname": "supplier_billing_state", "fieldtype": "Data", "width": 140},
		{"label": _("Supplier Name"), "fieldname": "supplier_name", "fieldtype": "Data", "width": 180},
		{"label": _("Is Reverse Charge Status as per PR"), "fieldname": "is_reverse_charge", "fieldtype": "Data", "width": 100},
	])

	# Collect IDs for bulk fetching
	gl_entries = []
	purchase_vouchers = {"Purchase Invoice": set(), "Purchase Receipt": set()}
	
	for row in data:
		# `name` is typically available on actual GL Entry rows (not grouping/totals)
		if row.get("name"):
			gl_entries.append(row.get("name"))
		
		v_type = row.get("voucher_type")
		v_no = row.get("voucher_no")
		if v_type in purchase_vouchers and v_no:
			purchase_vouchers[v_type].add(v_no)

	# Bulk fetch GL Entry creations in chunks to avoid large query string limits
	gl_creation_map = {}
	if gl_entries:
		for i in range(0, len(gl_entries), 1000):
			chunk = gl_entries[i:i+1000]
			res = frappe.db.get_all("GL Entry", filters={"name": ("in", chunk)}, fields=["name", "creation"], as_list=True)
			if res:
				gl_creation_map.update(dict(res))

	# Bulk fetch voucher details
	voucher_data = {}
	supplier_addresses = set()
	
	for v_type, v_nos in purchase_vouchers.items():
		if v_nos:
			v_nos_list = list(v_nos)
			for i in range(0, len(v_nos_list), 1000):
				chunk = v_nos_list[i:i+1000]
				fetched_data = frappe.db.get_all(
					v_type,
					filters={"name": ("in", chunk)},
					fields=[
						"name", "company_gstin", "billing_address_display", 
						"supplier_name", "supplier_address", "dispatch_address", "is_reverse_charge"
					]
				)
				for d in fetched_data:
					voucher_data.setdefault(v_type, {})[d.name] = d
					if d.get("supplier_address"):
						supplier_addresses.add(d.supplier_address)
					if d.get("dispatch_address"):
						supplier_addresses.add(d.dispatch_address)
					
	# Bulk fetch addresses for Supplier Billing State and GSTINs
	address_data = {}
	if supplier_addresses:
		addr_list = list(supplier_addresses)
		for i in range(0, len(addr_list), 1000):
			chunk = addr_list[i:i+1000]
			res = frappe.db.get_all("Address", filters={"name": ("in", chunk)}, fields=["name", "state", "gstin"])
			for r in res:
				address_data[r.name] = r

	# Merge data back into rows
	for row in data:
		# Map creation
		if row.get("name") and row.get("name") in gl_creation_map:
			row["creation"] = gl_creation_map[row["name"]]
			
		# Map Voucher details
		v_type = row.get("voucher_type")
		v_no = row.get("voucher_no")
		
		if v_type in voucher_data and v_no in voucher_data[v_type]:
			v_details = voucher_data[v_type][v_no]
			
			row["company_gstin"] = v_details.get("company_gstin")
			row["billing_address_display"] = v_details.get("billing_address_display")
			row["supplier_name"] = v_details.get("supplier_name")
			row["is_reverse_charge"] = "1" if v_details.get("is_reverse_charge") else "0"
			
			s_address = v_details.get("supplier_address")
			d_address = v_details.get("dispatch_address")
			
			if s_address and s_address in address_data:
				row["supplier_billing_state"] = address_data[s_address].get("state")
				row["supplier_billing_gstin"] = address_data[s_address].get("gstin")
				
			if d_address and d_address in address_data:
				row["supplier_shipping_gstin"] = address_data[d_address].get("gstin")

	return columns, data

@frappe.whitelist()
def download_colored_excel(filters=None):
    if isinstance(filters, str):
        filters = json.loads(filters)
        
    filters = frappe._dict(filters or {})
        
    columns, data = execute(filters)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "General Ledger Audit"
    
    # Write Headers
    for col_idx, col in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col.get("label"))
        cell.font = Font(bold=True)
        
    # Colors matching JS logic (light red background, red text)
    red_fill = PatternFill(start_color="fff1f1", end_color="fff1f1", fill_type="solid")
    red_font = Font(color="d9534f", bold=True)
    
    for row_idx, row in enumerate(data, start=2):
        posting_date = row.get("posting_date")
        creation = row.get("creation")
        
        apply_color = False
        
        if posting_date and creation:
            # Handle formats: posting_date can be date obj/string, creation can be datetime/string
            p_date_str = str(posting_date)[:10]
            c_date_str = str(creation)[:10]
            
            if p_date_str != c_date_str:
                apply_color = True
                
        for col_idx, col in enumerate(columns, start=1):
            fieldname = col.get("fieldname")
            val = row.get(fieldname)
            
            # Format val for excel properly if it's a date or datetime object, but str is usually fine for these reports
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            
            if apply_color:
                cell.fill = red_fill
                cell.font = red_font
                
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    # Output to response
    file_data = BytesIO()
    wb.save(file_data)
    
    frappe.response['filename'] = "General_Ledger_Audit.xlsx"
    frappe.response['filecontent'] = file_data.getvalue()
    frappe.response['type'] = 'binary'
